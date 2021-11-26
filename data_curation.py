import streamlit as st
import openpyxl
import re
import textblob
import autocorrect
import string
import pandas as pd
from openpyxl import Workbook
from autocorrect import Speller
import pickle
import os.path
from os import path
import time
import base64
import os


# pd.show_versions()
# inputPath = st.text_input(label='Enter the input excel file path')

st.title("Data curation")
File = st.file_uploader("Input excel file", type="xlsx")
# @st.cache()

if File is not None:
    st.markdown("File uploaded")
    # outputPath = st.text_input(label='Enter the output file path')
    wb_obj = openpyxl.load_workbook(File)
    sheet_obj = wb_obj.active

    # cell_obj = sheet_obj.cell(row=1,column=1)
    # sheet = wb_obj['Sheet0']
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column

    if sheet_obj.cell(row=1, column=1).value.upper() == "RECORDEDDATE":
        myList = []
        for i in range(3, max_row + 1):
            for j in range(2, max_col + 1):
                cell_obj = sheet_obj.cell(row=i, column=j)
                if cell_obj.value is not None:
                    myList.append(cell_obj.value.capitalize())
    else:
        myList = []
        for i in range(2, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=8)
            if cell_obj.value is not None and cell_obj.value != '':
                myList.append(cell_obj.value.capitalize())

    # print(myList)
    # print(len(myList))
    myList.sort()

    # print(myList)

    def remove_punc(string):
        punc = '''!()[]{};:'"\,<>./?@#$%^&*_~â€‹'''
        for ele in string:
            if ele in punc:
                string = string.replace(ele, "")
        return string


    myList = [remove_punc(i) for i in myList]

    # print(myList)

    def remove_duplicate(x):
        return list(dict.fromkeys(x))

    myList = remove_duplicate(myList)

    # print(myList)
    # print(len(myList))

    myWorkbook = openpyxl.Workbook()
    mySheet = myWorkbook.active

    mySheet["A1"] = "name"
    mySheet["B1"] = "content"

    base_path = os.path.dirname(os.getcwd())
    outputPath = base_path+'/SER'
    
    try:
#         directory = "SER"
#         parent_dir = "D:\\"
#         path = os.path.join(parent_dir, directory)
#         print(os.path.exists(path))
        # if os.path.exists(r"D:\SER"):
        #     outputPath = 'D:\\SER'
        # else:
        #     os.mkdir(r"D:\SER")
        #     outputPath = 'D:\\SER'
        does_path_exist = os.path.exists(outputPath)
        if not does_path_exist:
            os.mkdir(outputPath)

        #outputPath = 'C:'

        # outputPath = 'D:\\SERS\\5103 Indegenious\\'
        if outputPath:
            outCSVPath = outputPath

            # outCSVPath = 'D:/SERS/5103 Indegenious/'
            outputExcel = 'Processed takeaways' + '.xlsx'
            outputCSV = 'Processed takeaways' + '.csv'

            myWorkbook.save(outputPath + outputExcel)
        
            # spell check with autocorrect
            check = Speller(lang='en')
            cList = []

            for i in range(len(myList)):
                txt = myList[i]
                corrected = check(txt)
                cList.append(corrected)

            # print(cList)

            j = 0
            for i in range(2, len(myList) + 2):
                cellref = mySheet.cell(row=i, column=2)
                cellref.value = cList[j]
                j = j + 1

            k = 000
            # print(k)
            for i in range(2, len(myList) + 2):
                cellref = mySheet.cell(row=i, column=1)
                k = k + 1
                cellref.value = "A" + str(k)

            myWorkbook.save(outputPath + outputExcel)
        
            read_file = pd.read_excel(outCSVPath + outputExcel, sheet_name='Sheet')
        
            read_file.to_csv(outCSVPath + outputCSV, encoding='utf-8', index=None, header=True)
            # st.download_button(label="Download data as CSV", data=read_file, file_name='Processed takeaways.csv',mime='text/csv')

            def download_link_from_csv(csv, file_name, title="Download"):
                b64 = base64.b64encode(csv.encode()).decode()  # some strings <-> bytes conversions necessary here
                href = "<a href='data:file/csv;base64,{}' download='{}'>{}</a>".format(b64, file_name, title)
                st.markdown(href, unsafe_allow_html=True)


            timestr = time.strftime('%Y%m%d-%H%M%S')
            download_link_from_csv(read_file.to_csv(index=False,encoding='utf-8'), "Processed takeaways {}.csv".format(timestr), "Download processed takeaways")

            read_file.to_pickle('Processed.pickle')
        else:
            st.markdown("Error in output path")
    except OSError as error:
        st.markdown(error)
        st.markdown("Error during folder creation")

else:
    st.stop()
